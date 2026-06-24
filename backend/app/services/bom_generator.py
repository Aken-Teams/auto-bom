"""Generate output Excel files: C-CMAX import list + 3 ERP upload files."""
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2B3A67", end_color="2B3A67", fill_type="solid")
CELL_FONT = Font(name="Microsoft YaHei", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header_row(ws, headers, min_w=10, max_w=30):
    """Apply consistent header styling + sensible column widths + freeze the
    header row. Header TEXT is left untouched (must match ERP import format);
    this only adjusts formatting so columns are readable and not truncated.
    """
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        letter = get_column_letter(col)
        text = "" if header is None else str(header)
        longest = max((len(line) for line in text.split("\n")), default=0)
        ws.column_dimensions[letter].width = min(max(longest + 2, min_w), max_w)
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"


def _style_header(ws, headers: list[str]):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_row(ws, row_idx: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = CELL_FONT
        cell.border = THIN_BORDER


def _parse_component_summary(summary: str) -> dict:
    """Parse component_summary like 'SKY/5"/2SB229100PMDYY-250-K/90/*90mil/250um/AGAG/含砷'."""
    parts = (summary or "").split("/")
    return {
        "supplier_abbr": parts[0] if len(parts) > 0 else "",
        "wafer_size": parts[1] if len(parts) > 1 else "",
        "wafer_type": parts[2] if len(parts) > 2 else "",
        "mil_num": parts[3] if len(parts) > 3 else "",
        "mil_str": parts[4] if len(parts) > 4 else "",
        "thickness": parts[5] if len(parts) > 5 else "",
        "metal": parts[6] if len(parts) > 6 else "",
        "note": parts[7] if len(parts) > 7 else "",
    }


def _max_alt(alt_structure: str) -> str:
    """Derive MAX替代结构 from 替代结构 (append 'M' if not already present)."""
    if not alt_structure:
        return ""
    return alt_structure if alt_structure.endswith("M") else alt_structure + "M"


# ---------------------------------------------------------------------------
# 1. C-CMAX Import List
# ---------------------------------------------------------------------------

def generate_cmax_list(items: list[dict], output_dir: Path) -> str:
    """Generate C-CMAX import list Excel file (料号清单 + 罐头 sheets).

    Matches reference format: 20 columns on 料号清单, 16 columns on 罐头.
    """
    wb = Workbook()

    # --- Sheet 1: 料号清单 (20 columns) ---
    ws1 = wb.active
    ws1.title = "料号清单"
    headers1 = [
        "料号", "摘要", "内规文件编号", "大分类", "中分类",
        "替代结构", "MAX替代结构", "TYPE", "FAMILY", "PACKAGE",
        "LINE", "FUNCTION", "料号序号", "原件", "原件摘要",
        "晶片供应商", "Wafer Type", "原件附注", "单位用量", "原件生效时间",
    ]
    _style_header(ws1, headers1)

    for idx, item in enumerate(items, 2):
        alt = item.get("alt_structure", "")
        comp_summary = item.get("component_summary", "")
        parsed = _parse_component_summary(comp_summary)

        row_data = [
            item.get("item_no", ""),
            item.get("summary", ""),
            item.get("doc_no", ""),
            item.get("category_l", ""),
            item.get("category_m", ""),
            alt,
            _max_alt(alt),
            item.get("type_name", ""),
            item.get("family", ""),
            item.get("package", ""),
            item.get("line", ""),
            item.get("function", ""),
            item.get("seq_no", 10),
            item.get("component", ""),
            comp_summary,
            item.get("supplier", "") or "",          # 晶片供应商
            parsed.get("wafer_type", ""),             # Wafer Type
            item.get("bom_note", "") or "",           # 原件附注
            item.get("unit_usage", "") or "",         # 单位用量
            item.get("effective_date", "") or "",     # 原件生效时间
        ]
        for col, val in enumerate(row_data, 1):
            ws1.cell(row=idx, column=col, value=val)
        _style_row(ws1, idx, len(headers1))

    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[chr(64 + col) if col < 27 else "A" + chr(64 + col - 26)].width = 18

    # --- Sheet 2: 罐头 (16 columns matching reference) ---
    ws2 = wb.create_sheet("罐头")
    headers2 = [
        "A1", "A3", "A4", "A5", "A6", "A7", "A8", "A9", "A10", "A11",
        "焊接罐头", None, "成型罐头", None, "包装罐头", None,
    ]
    for col, header in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=header)

    seen_cans = set()
    can_row = 2
    for item in items:
        comp = item.get("component", "")
        if not comp or comp in seen_cans:
            continue
        seen_cans.add(comp)

        parsed = _parse_component_summary(item.get("component_summary", ""))
        try:
            mil_num = int(parsed["mil_num"]) if parsed["mil_num"] else ""
        except (ValueError, TypeError):
            mil_num = parsed["mil_num"]

        row_data = [
            item.get("function", ""),       # A1 = FUNCTION
            comp,                            # A3 = WAF code
            parsed["supplier_abbr"],         # A4 = supplier abbreviation
            parsed["wafer_size"],            # A5 = wafer size
            parsed["wafer_type"],            # A6 = wafer type
            mil_num,                         # A7 = mil number
            parsed["mil_str"],              # A8 = mil string
            parsed["thickness"],            # A9 = thickness
            parsed["metal"],                # A10 = metal
            parsed["note"],                 # A11 = note
            item.get("weld_can", ""),       # 焊接罐头 code
            item.get("weld_desc", ""),      # 焊接罐头 description
            item.get("mold_can", ""),       # 成型罐头 code
            item.get("mold_desc", ""),      # 成型罐头 description
            item.get("pack_can", ""),       # 包装罐头 code
            item.get("pack_desc", ""),      # 包装罐头 description
        ]
        for col, val in enumerate(row_data, 1):
            ws2.cell(row=can_row, column=col, value=val)
        can_row += 1

    for col in range(1, 17):
        ws2.column_dimensions[chr(64 + col)].width = 20

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = output_dir / f"C-CMAX_import_list_{ts}.xlsx"
    wb.save(str(filepath))
    return str(filepath)


# ---------------------------------------------------------------------------
# 2. pj_bom_loader
# ---------------------------------------------------------------------------

# Full 23-column header list matching reference
_BOM_HEADERS = [
    "",
    "assembly_item\n組裝料號料號",
    "alternate_bom_designator\n替代結構",
    "operation_sequence_number\n作業序號",
    "component_item\n元件",
    "quantity_per_assembly\nERP數量",
    "component_yield_factor\n良品率",
    "附註",
    None, None, None, None, None, None, None, None,   # cols 9-16 empty
    "BOP",
    "PROCESS_SPEC",
    "BOM_NAME",
    "CHIP_QTY\n晶粒數",
    "WIR_QTY\n打線數",
    "DB SEQUENCE\n晶片焊接 順序",
    "COMPONENT_ALTERNATE_QTY\nMES元件耗用量",
]


def generate_bom_loader(items: list[dict], output_dir: Path) -> str:
    """Generate pj_bom_loader Excel file (23 columns matching reference)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "主BOM"

    # Write headers
    for col, header in enumerate(_BOM_HEADERS, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, _BOM_HEADERS)

    row_idx = 2
    for item in items:
        item_no = item.get("item_no", "")
        alt = item.get("alt_structure", "")
        max_a = _max_alt(alt)
        component = item.get("component", "")
        weld = item.get("weld_can", "")
        mold = item.get("mold_can", "")
        pack = item.get("pack_can", "")
        package = item.get("package", "")

        bop = alt[:3] if alt else ""
        process_spec = f"{bop}_{package}(MAX)" if bop and package else ""
        bom_name = f"{item_no}_{max_a}" if item_no and max_a else item_no

        # BOM row template: [col2=item, col3=alt, col4=seq, col5=comp, col6=qty, col7=yield, ..., col17=BOP, col18=PSPEC, col19=BNAME]
        def _write_bom_row(seq, comp, qty=1):
            nonlocal row_idx
            ws.cell(row=row_idx, column=2, value=item_no)
            ws.cell(row=row_idx, column=3, value=max_a)
            ws.cell(row=row_idx, column=4, value=seq)
            ws.cell(row=row_idx, column=5, value=comp)
            ws.cell(row=row_idx, column=6, value=qty)
            ws.cell(row=row_idx, column=7, value=1)
            ws.cell(row=row_idx, column=17, value=bop)
            ws.cell(row=row_idx, column=18, value=process_spec)
            ws.cell(row=row_idx, column=19, value=bom_name)
            row_idx += 1

        # seq 10: 切割 - WAF component
        if component:
            _write_bom_row(10, component, 1)

        # seq 20: 焊接罐头
        if weld:
            _write_bom_row(20, weld)

        # seq 30: 成型罐头
        if mold:
            _write_bom_row(30, mold)

        # seq 80: 包装罐头
        if pack:
            _write_bom_row(80, pack)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = output_dir / f"pj_bom_loader_{ts}.xlsx"
    wb.save(str(filepath))
    return str(filepath)


# ---------------------------------------------------------------------------
# 3. Routings
# ---------------------------------------------------------------------------

# Full 60-column header list matching reference
_ROUTING_HEADERS = [
    "   ",  # col 1
    "ROUTING_SEQUENCE_ID", "ASSEMBLY_ITEM_ID", "ORGANIZATION_ID",
    "ALTERNATE_ROUTING_DESIGNATOR", "LAST_UPDATE_DATE", "LAST_UPDATED_BY",
    "CREATION_DATE", "CREATED_BY", "LAST_UPDATE_LOGIN", "ROUTING_TYPE",
    "COMMON_ASSEMBLY_ITEM_ID", "COMMON_ROUTING_SEQUENCE_ID", "ROUTING_COMMENT",
    "COMPLETION_SUBINVENTORY", "COMPLETION_LOCATOR_ID", "ATTRIBUTE_CATEGORY",
    "ATTRIBUTE1", "ATTRIBUTE2", "ATTRIBUTE3", "ATTRIBUTE4", "ATTRIBUTE5",
    "ATTRIBUTE6", "ATTRIBUTE7", "ATTRIBUTE8", "ATTRIBUTE9", "ATTRIBUTE10",
    "ATTRIBUTE11", "ATTRIBUTE12", "ATTRIBUTE13", "ATTRIBUTE14", "ATTRIBUTE15",
    "REQUEST_ID", "PROGRAM_APPLICATION_ID", "PROGRAM_ID", "PROGRAM_UPDATE_DATE",
    "DEMAND_SOURCE_LINE", "SET_ID", "PROCESS_REVISION",
    "DEMAND_SOURCE_TYPE", "DEMAND_SOURCE_HEADER_ID",
    "ORGANIZATION_CODE", "ASSEMBLY_ITEM_NUMBER", "COMMON_ITEM_NUMBER",
    "LOCATION_NAME", "TRANSACTION_ID", "PROCESS_FLAG", "TRANSACTION_TYPE",
    "LINE_ID", "LINE_CODE", "MIXED_MODEL_MAP_FLAG", "PRIORITY",
    "CFM_ROUTING_FLAG", "TOTAL_PRODUCT_CYCLE_TIME", "CTP_FLAG",
    "ORIGINAL_SYSTEM_REFERENCE", "SERIALIZATION_START_OP",
    "DELETE_GROUP_NAME", "DG_DESCRIPTION", "BATCH_ID",
]


def generate_routings(items: list[dict], output_dir: Path) -> str:
    """Generate routings Excel file (60 columns matching reference)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "替代结构"

    # Write headers
    for col, header in enumerate(_ROUTING_HEADERS, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, _ROUTING_HEADERS)

    seen = set()
    row_idx = 2
    for item in items:
        item_no = item.get("item_no", "")
        alt = item.get("alt_structure", "")
        max_a = _max_alt(alt)
        key = (item_no, max_a)
        if key in seen:
            continue
        seen.add(key)

        # Populated columns per reference:
        ws.cell(row=row_idx, column=5, value=max_a)          # ALTERNATE_ROUTING_DESIGNATOR
        ws.cell(row=row_idx, column=11, value=1)             # ROUTING_TYPE
        ws.cell(row=row_idx, column=40, value="0")           # DEMAND_SOURCE_TYPE
        ws.cell(row=row_idx, column=42, value="WX1")         # ORGANIZATION_CODE
        ws.cell(row=row_idx, column=43, value=item_no)       # ASSEMBLY_ITEM_NUMBER
        ws.cell(row=row_idx, column=47, value=1)             # PROCESS_FLAG
        ws.cell(row=row_idx, column=48, value="CREATE")      # TRANSACTION_TYPE
        row_idx += 1

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = output_dir / f"routings_{ts}.xlsx"
    wb.save(str(filepath))
    return str(filepath)


# ---------------------------------------------------------------------------
# 4. Sequences raw
# ---------------------------------------------------------------------------

# Full 99-column header list matching reference
_SEQ_HEADERS = [
    "   ",  # col 1
    "OPERATION_SEQUENCE_ID", "ROUTING_SEQUENCE_ID",
    "OPERATION_SEQ_NUM", "LAST_UPDATE_DATE", "LAST_UPDATED_BY",
    "CREATION_DATE", "CREATED_BY", "LAST_UPDATE_LOGIN",
    "STANDARD_OPERATION_ID", "DEPARTMENT_ID",
    "OPERATION_LEAD_TIME_PERCENT", "RUN_TIME_OVERLAP_PERCENT",
    "MINIMUM_TRANSFER_QUANTITY", "COUNT_POINT_TYPE",
    "OPERATION_DESCRIPTION", "EFFECTIVITY_DATE",
    "CHANGE_NOTICE", "IMPLEMENTATION_DATE", "DISABLE_DATE",
    "BACKFLUSH_FLAG", "OPTION_DEPENDENT_FLAG", "ATTRIBUTE_CATEGORY",
    "ATTRIBUTE1", "ATTRIBUTE2", "ATTRIBUTE3", "ATTRIBUTE4", "ATTRIBUTE5",
    "ATTRIBUTE6", "ATTRIBUTE7", "ATTRIBUTE8", "ATTRIBUTE9", "ATTRIBUTE10",
    "ATTRIBUTE11", "ATTRIBUTE12", "ATTRIBUTE13", "ATTRIBUTE14", "ATTRIBUTE15",
    "REQUEST_ID", "PROGRAM_APPLICATION_ID", "PROGRAM_ID", "PROGRAM_UPDATE_DATE",
    "MODEL_OP_SEQ_ID", "ASSEMBLY_ITEM_ID", "ORGANIZATION_ID",
    "ALTERNATE_ROUTING_DESIGNATOR", "ORGANIZATION_CODE", "ASSEMBLY_ITEM_NUMBER",
    "DEPARTMENT_CODE", "OPERATION_CODE",
    "RESOURCE_ID1", "RESOURCE_ID2", "RESOURCE_ID3",
    "RESOURCE_CODE1", "RESOURCE_CODE2", "RESOURCE_CODE3",
    "INSTRUCTION_CODE1", "INSTRUCTION_CODE2", "INSTRUCTION_CODE3",
    "TRANSACTION_ID", "PROCESS_FLAG", "TRANSACTION_TYPE",
    "NEW_OPERATION_SEQ_NUM", "NEW_EFFECTIVITY_DATE",
    "ASSEMBLY_TYPE", "OPERATION_TYPE", "REFERENCE_FLAG",
    "PROCESS_OP_SEQ_ID", "LINE_OP_SEQ_ID",
    "YIELD", "CUMULATIVE_YIELD", "REVERSE_CUMULATIVE_YIELD",
    "LABOR_TIME_CALC", "MACHINE_TIME_CALC", "TOTAL_TIME_CALC",
    "LABOR_TIME_USER", "MACHINE_TIME_USER", "TOTAL_TIME_USER",
    "NET_PLANNING_PERCENT", "INCLUDE_IN_ROLLUP", "OPERATION_YIELD_ENABLED",
    "PROCESS_SEQ_NUMBER", "PROCESS_CODE",
    "LINE_OP_SEQ_NUMBER", "LINE_OP_CODE",
    "ORIGINAL_SYSTEM_REFERENCE", "SHUTDOWN_TYPE", "LONG_DESCRIPTION",
    "DELETE_GROUP_NAME", "DG_DESCRIPTION",
    "NEW_ROUTING_REVISION", "ACD_TYPE", "OLD_START_EFFECTIVE_DATE",
    "CANCEL_COMMENTS", "ENG_CHANGES_IFCE_KEY", "ENG_REVISED_ITEMS_IFCE_KEY",
    "BOM_REV_OP_IFCE_KEY", "NEW_REVISED_ITEM_REVISION", "BATCH_ID",
]

# Fixed sequence operations: seq_num -> department_code
_SEQ_STEPS = [
    (10, "切割"),
    (20, "焊接"),
    (30, "成型"),
    (40, "切脚"),
    (50, "Burning"),
    (60, "外包前"),
    (61, "贝维特"),
    (63, "佰润"),
    (68, "欣捷"),
    (70, "外包后"),
    (80, "TMTT"),
    (90, "FQC"),
]


def generate_sequences(items: list[dict], std_ops: dict, output_dir: Path) -> str:
    """Generate sequences-raw Excel file (99 columns, 12 operations per item).

    std_ops: dict mapping (seq_num, summary_keyword) -> standard_operation_id
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "替代结构"

    # Write headers
    for col, header in enumerate(_SEQ_HEADERS, 1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, _SEQ_HEADERS)

    now = datetime.now()
    seen = set()
    row_idx = 2
    for item in items:
        item_no = item.get("item_no", "")
        alt = item.get("alt_structure", "")
        max_a = _max_alt(alt)
        key = (item_no, max_a)
        if key in seen:
            continue
        seen.add(key)

        for seq_num, dept_code in _SEQ_STEPS:
            # Lookup standard operation ID from std_ops dict
            std_op_id = std_ops.get(seq_num, "")

            ws.cell(row=row_idx, column=4, value=seq_num)                        # OPERATION_SEQ_NUM
            if std_op_id:
                ws.cell(row=row_idx, column=10, value=std_op_id)                 # STANDARD_OPERATION_ID
            ws.cell(row=row_idx, column=16, value=f"sequence_{seq_num}")          # OPERATION_DESCRIPTION
            ws.cell(row=row_idx, column=17, value=now)                            # EFFECTIVITY_DATE
            ws.cell(row=row_idx, column=46, value=max_a)                          # ALTERNATE_ROUTING_DESIGNATOR
            ws.cell(row=row_idx, column=47, value="WX1")                          # ORGANIZATION_CODE
            ws.cell(row=row_idx, column=48, value=item_no)                        # ASSEMBLY_ITEM_NUMBER
            ws.cell(row=row_idx, column=49, value=dept_code)                      # DEPARTMENT_CODE
            ws.cell(row=row_idx, column=61, value=1)                              # PROCESS_FLAG
            ws.cell(row=row_idx, column=62, value="CREATE")                       # TRANSACTION_TYPE
            ws.cell(row=row_idx, column=67, value=1)                              # REFERENCE_FLAG
            row_idx += 1

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = output_dir / f"sequences_raw_{ts}.xlsx"
    wb.save(str(filepath))
    return str(filepath)
