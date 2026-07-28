"""Parse uploaded Excel files (BOM base, standard operations)."""
from datetime import datetime, time
from pathlib import Path
from openpyxl import load_workbook


def _unit_usage(v):
    """单位用量 may be date-formatted (old template: 1 -> datetime(1900,1,1),
    0 -> time(0,0)) or a plain integer (fixed template 7-23: 1 / 0)."""
    if isinstance(v, datetime):
        return 1
    if isinstance(v, time):
        return 0
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return int(v)
    return None


# --- Header-driven column mapping -------------------------------------------
# Templates have varied column ORDER over time (WXBMR005 底稿, C-CMAX 导入清单,
# 固定模板 7-23). Rather than read by fixed position — which silently mis-reads
# when a column moves — locate each field by its header NAME. Position is used
# only as a legacy fallback for headerless sheets.
_FIELD_ALIASES = {
    "item_no": ["料号", "料號"],
    "summary": ["摘要"],
    "doc_no": ["内规文件编号", "內規文件編號"],
    "category_l": ["大分类", "大分類"],
    "category_m": ["中分类", "中分類"],
    "alt_structure": ["替代结构", "替代結構"],
    "max_alt_structure": ["MAX替代结构", "MAX替代結構"],
    "type_name": ["TYPE"],
    "family": ["FAMILY"],
    "package": ["PACKAGE"],
    "line": ["LINE"],
    "function": ["FUNCTION"],
    "seq_no": ["料号序号", "料號序號"],
    "component": ["原件"],
    "component_summary": ["原件摘要"],
    "unit_usage": ["单位用量", "單位用量"],
}
# Legacy positional layout (WXBMR005 / C-CMAX) used only when no header row.
_LEGACY_POS = {
    "item_no": 0, "summary": 1, "doc_no": 2, "category_l": 3, "category_m": 4,
    "alt_structure": 5, "max_alt_structure": 6, "type_name": 7, "family": 8,
    "package": 9, "line": 10, "function": 11, "seq_no": 12, "component": 13,
    "component_summary": 14, "unit_usage": 18,
}


def _header_index(ws) -> dict:
    """Map each (stripped, exact) header name in row 1 to its column index."""
    idx = {}
    for j, cell in enumerate(ws[1]):
        if cell.value is None:
            continue
        name = str(cell.value).strip()
        if name and name not in idx:
            idx[name] = j
    return idx


def _read_item_rows(ws) -> list[dict]:
    """Read 料号 rows from a worksheet, locating columns by header name.

    Works across WXBMR005 底稿, C-CMAX 导入清单 and 固定模板 7-23 regardless of
    column order. Falls back to legacy fixed positions only if there is no
    recognizable header row (no 料号 column)."""
    hdr = _header_index(ws)
    header_mode = "料号" in hdr or "料號" in hdr

    def col_of(field):
        if header_mode:
            for alias in _FIELD_ALIASES[field]:
                if alias in hdr:
                    return hdr[alias]
            return None  # optional field absent in this template
        return _LEGACY_POS[field]

    cols = {f: col_of(f) for f in _FIELD_ALIASES}

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)

        def get(field):
            i = cols[field]
            if i is None or i >= len(vals):
                return None
            return vals[i]

        if not get("item_no"):
            continue
        items.append({
            "item_no": str(get("item_no") or ""),
            "summary": str(get("summary") or ""),
            "doc_no": str(get("doc_no") or ""),
            "category_l": str(get("category_l") or ""),
            "category_m": str(get("category_m") or ""),
            "alt_structure": str(get("alt_structure") or ""),
            "max_alt_structure": str(get("max_alt_structure") or ""),
            "bom_note": "",
            "type_name": str(get("type_name") or ""),
            "family": str(get("family") or ""),
            "package": str(get("package") or ""),
            "line": str(get("line") or ""),
            "function": str(get("function") or ""),
            "seq_no": str(get("seq_no") or ""),
            "component": str(get("component") or ""),
            "component_summary": str(get("component_summary") or ""),
            "unit_usage": _unit_usage(get("unit_usage")),
        })
    return items


def parse_bom_base(file_path: str) -> list[dict]:
    """Parse MBU2 BOM底稿 (WXBMR005) / 导入清单 / 固定模板 — header-driven."""
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    items = _read_item_rows(ws)
    wb.close()
    return items


def _can_label(desc: str) -> str:
    """Build a short display label from a can description.

    e.g. "包裝/SMC/TR 7\"/////0.8K/" -> "SMC / TR 7\" / 0.8K"
    Drops the leading type segment and any empty segments.
    """
    parts = [p.strip() for p in desc.split("/")]
    meaningful = [p for p in parts[1:] if p]
    return " / ".join(meaningful)


def _weld_prefix_mil(desc: str) -> tuple[str, str]:
    """Extract the切割/焊接前缀 and mil from a weld description.

    e.g. "焊接/SMC-MAX/ACP/C(C)/C(E)//SKY_55MIL//" -> ("SKY", "55")
         "焊接/.../EURG_70MIL//" -> ("EURG", "70")
    """
    import re
    m = re.search(r"([A-Za-z]+)_(\d+(?:\.\d+)?)MIL", desc or "")
    return (m.group(1), m.group(2)) if m else ("", "")


def _parse_can_fixed(ws) -> dict:
    """Parse the 固定模板 7-23 罐头 layout: three parallel column pairs
    焊接罐头 | 摘要 | 成型罐头 | 摘要 | 包装罐头 | 摘要.

    Unlike the legacy layout this carries NO WAF / function / mil columns — the
    weld's prefix + mil live inside the description (SKY_55MIL), so we parse them
    out and store the prefix in `function` (waf_code stays empty)."""
    hdr = _header_index(ws)

    def pair(name):
        i = hdr.get(name)
        return (i, i + 1) if i is not None else (None, None)

    wc, wd = pair("焊接罐头")
    mc, md = pair("成型罐头")
    pc, pd = pair("包装罐头")

    weld_templates, options, seen = [], [], set()

    def cell(vals, i):
        return str(vals[i] or "") if i is not None and i < len(vals) else ""

    def add_option(can_type, code, desc):
        key = (can_type, code)
        if code and key not in seen:
            seen.add(key)
            options.append({"can_type": can_type, "can_code": code,
                            "can_desc": desc, "label": _can_label(desc)})

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        wcode, wdesc = cell(vals, wc), cell(vals, wd)
        if wcode:
            prefix, mil = _weld_prefix_mil(wdesc)
            weld_templates.append({
                "function": prefix, "waf_code": "", "supplier": "",
                "wafer_size": "", "mil": mil, "weld_can": wcode,
                "weld_desc": wdesc, "mold_can": "", "mold_desc": "",
                "pack_can": "", "pack_desc": "",
            })
        add_option("mold", cell(vals, mc), cell(vals, md))
        add_option("pack", cell(vals, pc), cell(vals, pd))

    return {"weld": weld_templates, "options": options}


def parse_can_template(file_path: str) -> dict:
    """Parse a 罐头 sheet.

    Two layouts are supported:
      * Legacy (C-CMAX): one can per row, type encoded in the description
        (焊接 -> weld matched by WAF; 成型 -> mold; 包裝/包装 -> pack).
      * 固定模板 7-23: three parallel column pairs
        (焊接罐头|摘要|成型罐头|摘要|包装罐头|摘要) — see _parse_can_fixed.

    Returns {"weld": [<CanTemplate dicts>], "options": [<general can dicts>]}.
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb["罐头"] if "罐头" in wb.sheetnames else wb.active

    if "焊接罐头" in _header_index(ws):  # 固定模板 layout
        result = _parse_can_fixed(ws)
        wb.close()
        return result

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    weld_templates = []
    options = []
    seen_options = set()
    for row in rows:
        vals = list(row)
        waf = str(vals[1] or "") if len(vals) > 1 else ""
        can_code = str(vals[10] or "") if len(vals) > 10 else ""
        can_desc = str(vals[11] or "") if len(vals) > 11 else ""
        if not can_code and not waf:
            continue

        if can_desc.startswith("成型"):
            can_type = "mold"
        elif can_desc.startswith("包裝") or can_desc.startswith("包装"):
            can_type = "pack"
        elif can_desc.startswith("焊接") or waf:
            can_type = "weld"
        else:
            continue

        if can_type == "weld":
            weld_templates.append({
                "function": str(vals[0] or ""),
                "waf_code": waf,
                "supplier": str(vals[2] or "") if len(vals) > 2 else "",
                "wafer_size": str(vals[3] or "") if len(vals) > 3 else "",
                "mil": str(vals[5] or "") if len(vals) > 5 else "",
                "weld_can": can_code,
                "weld_desc": can_desc,
                "mold_can": "",
                "mold_desc": "",
                "pack_can": "",
                "pack_desc": "",
            })
        else:
            key = (can_type, can_code)
            if can_code and key not in seen_options:
                seen_options.add(key)
                options.append({
                    "can_type": can_type,
                    "can_code": can_code,
                    "can_desc": can_desc,
                    "label": _can_label(can_desc),
                })

    return {"weld": weld_templates, "options": options}


def parse_item_list(file_path: str) -> list[dict]:
    """Parse 料号清单 (C-CMAX导入清单 / 固定模板) — header-driven."""
    wb = load_workbook(file_path, data_only=True)
    ws = wb["料号清单"] if "料号清单" in wb.sheetnames else wb.active
    items = _read_item_rows(ws)
    wb.close()
    return items


def parse_std_operations(file_path: str, limit: int = 0) -> list[dict]:
    """Parse WXBMR004 standard operations.

    Args:
        file_path: Path to the Excel file.
        limit: If > 0, only parse the first `limit` valid rows (for quick validation).
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    ops = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if not vals[0]:
            continue
        try:
            op_id = int(vals[0])
        except (ValueError, TypeError):
            continue  # skip non-numeric rows (wrong file or header)
        try:
            seq = int(vals[5]) if vals[5] else 0
        except (ValueError, TypeError):
            seq = 0
        ops.append({
            "op_id": op_id,
            "code": str(vals[1] or ""),
            "summary": str(vals[2] or ""),
            "department": str(vals[3] or ""),
            "dept_summary": str(vals[4] or ""),
            "seq": seq,
            "resource": str(vals[6] or "") if len(vals) > 6 else "",
            "resource_summary": str(vals[7] or "") if len(vals) > 7 else "",
            "unit": str(vals[8] or "") if len(vals) > 8 else "",
            "pph": str(vals[9] or "") if len(vals) > 9 else "",
        })
        if limit and len(ops) >= limit:
            break
    wb.close()
    return ops
