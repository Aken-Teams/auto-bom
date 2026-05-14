"""Compare generated output files against reference data files."""
import openpyxl
import sys
import io
from pathlib import Path

# Force UTF-8 output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')


def load_sheet_data(filepath, sheet_name=None, max_rows=None):
    """Load sheet data from an Excel file. Returns (headers, rows, sheet_names)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_names = wb.sheetnames

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            return None, None, sheet_names
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [], sheet_names

    headers = list(rows[0])
    data_rows = rows[1:]
    if max_rows is not None:
        data_rows = data_rows[:max_rows]

    wb.close()
    return headers, data_rows, sheet_names


def normalize_header(h):
    """Normalize header for comparison - strip whitespace, collapse newlines."""
    if h is None:
        return ""
    return str(h).strip().replace("\r\n", "\n").replace("\r", "\n")


def compare_headers(ref_headers, gen_headers, label):
    """Compare two header lists and report differences."""
    ref_norm = [normalize_header(h) for h in ref_headers]
    gen_norm = [normalize_header(h) for h in gen_headers]

    # Remove trailing empty headers
    while ref_norm and ref_norm[-1] == "":
        ref_norm.pop()
    while gen_norm and gen_norm[-1] == "":
        gen_norm.pop()

    issues = []

    if len(ref_norm) != len(gen_norm):
        issues.append(f"  Column count differs: reference={len(ref_norm)}, generated={len(gen_norm)}")

    # Find columns in reference but missing in generated
    ref_set = set(ref_norm)
    gen_set = set(gen_norm)
    missing = ref_set - gen_set
    extra = gen_set - ref_set

    if missing:
        issues.append(f"  MISSING columns (in reference but not in generated):")
        for m in sorted(missing):
            issues.append(f"    - '{m}'")

    if extra:
        issues.append(f"  EXTRA columns (in generated but not in reference):")
        for e in sorted(extra):
            issues.append(f"    - '{e}'")

    # Check order for matching columns
    min_len = min(len(ref_norm), len(gen_norm))
    order_diffs = []
    for i in range(min_len):
        if ref_norm[i] != gen_norm[i]:
            order_diffs.append(f"    Col {i+1}: ref='{ref_norm[i]}' vs gen='{gen_norm[i]}'")

    if order_diffs:
        issues.append(f"  Column ORDER differences:")
        for d in order_diffs:
            issues.append(d)

    return issues, ref_norm, gen_norm


def format_cell(val):
    """Format cell value for display."""
    if val is None:
        return "<empty>"
    return repr(val)


def compare_sample_rows(ref_rows, gen_rows, ref_headers, gen_headers, max_show=5):
    """Compare sample data rows."""
    issues = []

    ref_count = len(ref_rows)
    gen_count = len(gen_rows)

    issues.append(f"  Data rows: reference={ref_count}, generated={gen_count}")

    # Show first few rows from each
    issues.append(f"\n  --- First {min(max_show, ref_count)} rows from REFERENCE ---")
    for i, row in enumerate(ref_rows[:max_show]):
        row_display = []
        for j, val in enumerate(row):
            if val is not None and str(val).strip() != "":
                hdr = ref_headers[j] if j < len(ref_headers) else f"Col{j+1}"
                # Truncate long header names for display
                hdr_short = hdr.split("\n")[0][:20]
                row_display.append(f"{hdr_short}={format_cell(val)}")
        issues.append(f"    Row {i+1}: {', '.join(row_display)}")

    issues.append(f"\n  --- First {min(max_show, gen_count)} rows from GENERATED ---")
    for i, row in enumerate(gen_rows[:max_show]):
        row_display = []
        for j, val in enumerate(row):
            if val is not None and str(val).strip() != "":
                hdr = gen_headers[j] if j < len(gen_headers) else f"Col{j+1}"
                hdr_short = hdr.split("\n")[0][:20]
                row_display.append(f"{hdr_short}={format_cell(val)}")
        issues.append(f"    Row {i+1}: {', '.join(row_display)}")

    return issues


def compare_files(ref_path, gen_path, file_label):
    """Compare a reference file with a generated file."""
    print(f"\n{'='*80}")
    print(f"COMPARING: {file_label}")
    print(f"  Reference: {ref_path.name}")
    print(f"  Generated: {gen_path.name}")
    print(f"{'='*80}")

    ref_wb = openpyxl.load_workbook(ref_path, data_only=True)
    gen_wb = openpyxl.load_workbook(gen_path, data_only=True)

    ref_sheets = ref_wb.sheetnames
    gen_sheets = gen_wb.sheetnames

    print(f"\n  Sheet names:")
    print(f"    Reference: {ref_sheets}")
    print(f"    Generated: {gen_sheets}")

    ref_sheet_set = set(ref_sheets)
    gen_sheet_set = set(gen_sheets)

    missing_sheets = ref_sheet_set - gen_sheet_set
    extra_sheets = gen_sheet_set - ref_sheet_set

    if missing_sheets:
        print(f"    MISSING sheets: {missing_sheets}")
    if extra_sheets:
        print(f"    EXTRA sheets: {extra_sheets}")
    if not missing_sheets and not extra_sheets and ref_sheets == gen_sheets:
        print(f"    MATCH: Sheet names and order match")

    # Compare each sheet that exists in both
    all_sheets = list(dict.fromkeys(ref_sheets + gen_sheets))  # preserve order
    for sheet_name in all_sheets:
        print(f"\n  --- Sheet: '{sheet_name}' ---")

        if sheet_name not in ref_sheet_set:
            print(f"    Only in generated file (EXTRA)")
            # Still show its structure
            ws = gen_wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [normalize_header(h) for h in rows[0]]
                print(f"    Headers: {headers}")
                print(f"    Data rows: {len(rows) - 1}")
            continue

        if sheet_name not in gen_sheet_set:
            print(f"    Only in reference file (MISSING from generated)")
            ws = ref_wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [normalize_header(h) for h in rows[0]]
                print(f"    Headers: {headers}")
                print(f"    Data rows: {len(rows) - 1}")
            continue

        # Both exist - compare
        ref_ws = ref_wb[sheet_name]
        gen_ws = gen_wb[sheet_name]

        ref_rows = list(ref_ws.iter_rows(values_only=True))
        gen_rows = list(gen_ws.iter_rows(values_only=True))

        if not ref_rows:
            print(f"    Reference sheet is empty")
            continue
        if not gen_rows:
            print(f"    Generated sheet is empty")
            continue

        ref_headers = [normalize_header(h) for h in ref_rows[0]]
        gen_headers = [normalize_header(h) for h in gen_rows[0]]

        # Compare headers
        header_issues, ref_h_clean, gen_h_clean = compare_headers(
            ref_rows[0], gen_rows[0], f"{file_label}/{sheet_name}"
        )

        if header_issues:
            print(f"\n  Header comparison:")
            for issue in header_issues:
                print(issue)
        else:
            print(f"    Headers MATCH perfectly ({len(ref_h_clean)} columns)")

        # Print full header lists side by side
        print(f"\n  Full header listing:")
        max_cols = max(len(ref_headers), len(gen_headers))
        for i in range(max_cols):
            r = ref_headers[i] if i < len(ref_headers) else "<n/a>"
            g = gen_headers[i] if i < len(gen_headers) else "<n/a>"
            match_mark = "OK" if r == g else "DIFF"
            # Replace newlines for display
            r_display = r.replace("\n", " | ")
            g_display = g.replace("\n", " | ")
            print(f"    Col {i+1:2d}: [{match_mark:4s}] ref='{r_display}' vs gen='{g_display}'")

        # Compare row counts and sample data
        ref_data = ref_rows[1:]
        gen_data = gen_rows[1:]
        sample_issues = compare_sample_rows(ref_data, gen_data, ref_headers, gen_headers)
        for issue in sample_issues:
            print(issue)

        # Check data types in first few rows
        print(f"\n  Data type check (first 3 rows):")
        for row_i in range(min(3, len(ref_data))):
            type_diffs = []
            for col_i in range(min(len(ref_headers), len(ref_data[row_i]))):
                val = ref_data[row_i][col_i]
                if val is not None:
                    type_diffs.append(f"{ref_headers[col_i].split(chr(10))[0][:15]}={type(val).__name__}")
            if type_diffs:
                print(f"    Ref row {row_i+1} types: {', '.join(type_diffs[:10])}")

        if gen_data:
            for row_i in range(min(3, len(gen_data))):
                type_diffs = []
                for col_i in range(min(len(gen_headers), len(gen_data[row_i]))):
                    val = gen_data[row_i][col_i]
                    if val is not None:
                        type_diffs.append(f"{gen_headers[col_i].split(chr(10))[0][:15]}={type(val).__name__}")
                if type_diffs:
                    print(f"    Gen row {row_i+1} types: {', '.join(type_diffs[:10])}")

    ref_wb.close()
    gen_wb.close()


def main():
    data_dir = Path("D:/github/auto-bom/data")
    out_dir = Path("D:/github/auto-bom/outputs/task_5")

    comparisons = [
        (
            data_dir / "C-CMAX导入清单3-3(1).xlsx",
            out_dir / "C-CMAX_import_list_20260514_192857.xlsx",
            "C-CMAX Import List",
        ),
        (
            data_dir / "pj_bom_loader 3.3.xlsx",
            out_dir / "pj_bom_loader_20260514_192908.xlsx",
            "BOM Loader",
        ),
        (
            data_dir / "rountings 3.3.xlsx",
            out_dir / "routings_20260514_192908.xlsx",
            "Routings",
        ),
        (
            data_dir / "sequences-raw 3.3.xlsx",
            out_dir / "sequences_raw_20260514_192909.xlsx",
            "Sequences Raw",
        ),
    ]

    for ref_path, gen_path, label in comparisons:
        if not ref_path.exists():
            print(f"\nERROR: Reference file not found: {ref_path}")
            continue
        if not gen_path.exists():
            print(f"\nERROR: Generated file not found: {gen_path}")
            continue
        compare_files(ref_path, gen_path, label)

    print(f"\n{'='*80}")
    print("COMPARISON COMPLETE")
    print(f"{'='*80}")


if __name__ == "__main__":
    main()
